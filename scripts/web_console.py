#!/usr/bin/env python3
"""Lightweight console for district-form automation (year/quarter + pipeline)."""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import threading
from datetime import datetime
from io import BytesIO
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request
from werkzeug.utils import secure_filename

SCRIPT_DIR = Path(__file__).resolve().parent
PLUGIN_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(PLUGIN_ROOT))
load_dotenv(PLUGIN_ROOT / ".env")

from crawl_district_reports import load_settings, save_settings  # noqa: E402
from dfmc_browser_utils import (  # noqa: E402
    get_default_state_file,
    get_session_home,
    process_is_running,
    read_browser_state,
)
from excel_download import send_xlsx_file  # noqa: E402
from time_utils import beijing_strftime, ensure_beijing_tz  # noqa: E402

ensure_beijing_tz()

app = Flask(
    __name__,
    template_folder=str(PLUGIN_ROOT / "templates"),
)
DEFAULT_TEMPLATE_PATH = PLUGIN_ROOT / "app" / "assets" / "report_template.xlsx"
TEMPLATE_BACKUP_DIR = PLUGIN_ROOT / "app" / "assets" / "template_backups"

_pipeline_lock = threading.Lock()
_pipeline_state = {
    "running": False,
    "started_at": "",
    "finished_at": "",
    "status": "idle",
    "log": "",
    "result": {},
}


def _browser_ok() -> tuple[bool, dict]:
    state_file = get_default_state_file(PLUGIN_ROOT)
    if not state_file.exists():
        return False, {}
    try:
        state = read_browser_state(state_file)
    except Exception:
        return False, {}
    pid = int(state.get("pid") or 0)
    port = int(state.get("port") or 0)
    ok = bool(pid and process_is_running(pid) and port)
    return ok, state


def _latest_report() -> str:
    """Latest formula workbook (prefer legacy *_公式版.xlsx sidecar when present)."""
    output_dir = PLUGIN_ROOT / "output"
    if not output_dir.exists():
        return ""
    files = [
        p
        for p in output_dir.glob("区域各指标情况一览*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not files:
        return ""
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
    latest = files[0]
    # Legacy pipeline wrote values as main + *_公式版 sidecar; prefer the formula copy.
    if "_公式版" not in latest.name:
        companion = latest.with_name(f"{latest.stem}_公式版.xlsx")
        if companion.exists():
            return str(companion)
    return str(latest)


def _latest_manifest_text() -> str:
    path = PLUGIN_ROOT / "output" / "pipeline_manifest.json"
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def _template_meta_path() -> Path:
    return DEFAULT_TEMPLATE_PATH.with_suffix(".xlsx.meta.json")


def _template_info() -> dict:
    path = DEFAULT_TEMPLATE_PATH
    meta = {}
    meta_path = _template_meta_path()
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    if not path.exists():
        return {
            "exists": False,
            "path": str(path),
            "size": 0,
            "mtime": "",
            "uploaded_as": meta.get("uploaded_as") or "",
        }
    st = path.stat()
    return {
        "exists": True,
        "path": str(path),
        "size": st.st_size,
        "mtime": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "uploaded_as": meta.get("uploaded_as") or "",
        "uploaded_at": meta.get("uploaded_at") or "",
    }


def _backup_current_template() -> str:
    if not DEFAULT_TEMPLATE_PATH.exists():
        return ""
    TEMPLATE_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = beijing_strftime("%Y%m%d_%H%M%S")
    backup = TEMPLATE_BACKUP_DIR / f"report_template_{stamp}.xlsx"
    backup.write_bytes(DEFAULT_TEMPLATE_PATH.read_bytes())
    return str(backup)


def _validate_xlsx(content: bytes) -> None:
    if len(content) < 4 or content[:2] != b"PK":
        raise ValueError("不是有效的 .xlsx 文件（需为 Excel 工作簿）")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content), read_only=True, data_only=False)
        try:
            if not wb.sheetnames:
                raise ValueError("模板没有工作表")
        finally:
            wb.close()
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"模板无法打开：{exc}") from exc


def _parse_notify_flag(data: dict, *, default: bool | None = None) -> bool | None:
    if "notify_feishu" not in data and "skip_notify" not in data:
        return default
    if "notify_feishu" in data:
        return bool(data.get("notify_feishu"))
    return not bool(data.get("skip_notify"))


def _run_pipeline_job(
    year: int,
    quarter: int,
    skip_crawl: bool,
    *,
    notify_feishu: bool,
) -> None:
    global _pipeline_state
    log_path = PLUGIN_ROOT / "logs" / f"pipeline_{beijing_strftime('%Y%m%d_%H%M%S')}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    cmd = [
        str(PLUGIN_ROOT / ".venv" / "bin" / "python"),
        str(SCRIPT_DIR / "run_pipeline.py"),
        "--year",
        str(year),
        "--quarter",
        str(quarter),
        "--save-settings",
    ]
    if skip_crawl:
        cmd.append("--skip-crawl")
    if notify_feishu:
        cmd.append("--notify")
    else:
        cmd.append("--skip-notify")

    with _pipeline_lock:
        _pipeline_state.update({
            "running": True,
            "started_at": beijing_strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": "",
            "status": "running",
            "log": f"$ {' '.join(cmd)}\n",
            "result": {},
        })

    try:
        with log_path.open("w", encoding="utf-8") as fh:
            proc = subprocess.run(
                cmd,
                cwd=str(PLUGIN_ROOT),
                stdout=fh,
                stderr=subprocess.STDOUT,
                text=True,
                env={**os.environ, "TZ": "Asia/Shanghai"},
            )
        text = log_path.read_text(encoding="utf-8")
        status = "ok" if proc.returncode == 0 else f"exit_{proc.returncode}"
    except Exception as exc:
        text = str(exc)
        status = "error"

    result = {}
    manifest = PLUGIN_ROOT / "output" / "pipeline_manifest.json"
    if manifest.exists():
        try:
            result = json.loads(manifest.read_text(encoding="utf-8"))
        except Exception:
            result = {}

    with _pipeline_lock:
        _pipeline_state.update({
            "running": False,
            "finished_at": beijing_strftime("%Y-%m-%d %H:%M:%S"),
            "status": status,
            "log": text[-12000:],
            "result": result,
        })


@app.get("/")
def index():
    settings = load_settings()
    ok, state = _browser_ok()
    return render_template(
        "dashboard.html",
        session_home=str(get_session_home(PLUGIN_ROOT)),
        browser_ok=ok,
        browser_label=("在线" if ok else "离线") + (f" :{state.get('port')}" if ok else ""),
        year=int(settings["year"]),
        quarter=int(settings["quarter"]),
        notify_feishu=bool(settings.get("notify_feishu", True)),
        latest_report=_latest_report(),
    )


@app.get("/api/status")
def api_status():
    ok, state = _browser_ok()
    with _pipeline_lock:
        pipeline = dict(_pipeline_state)
    return jsonify({
        "browser_ok": ok,
        "browser": state,
        "session_home": str(get_session_home(PLUGIN_ROOT)),
        "settings": load_settings(),
        "pipeline": pipeline,
        "latest_report": _latest_report(),
        "latest_manifest": _latest_manifest_text(),
        "template": _template_info(),
    })


@app.post("/api/settings")
def api_settings():
    data = request.get_json(force=True, silent=True) or {}
    year = int(data.get("year") or 0)
    quarter = int(data.get("quarter") or 0)
    if year < 2020 or quarter not in {1, 2, 3, 4}:
        return jsonify({"error": "年份/季度无效"}), 400
    notify = _parse_notify_flag(data, default=None)
    save_settings(year, quarter, notify_feishu=notify)
    settings = load_settings()
    notify_text = "推送" if settings.get("notify_feishu") else "不推送"
    return jsonify({
        "message": f"已保存 {year} Q{quarter}，飞书{notify_text}",
        "settings": settings,
    })


@app.get("/api/template")
def api_template_info():
    return jsonify(_template_info())


@app.get("/api/template/download")
def api_template_download():
    if not DEFAULT_TEMPLATE_PATH.exists():
        return jsonify({"error": "默认模板不存在"}), 404
    return send_xlsx_file(DEFAULT_TEMPLATE_PATH, display_name="report_template.xlsx")


@app.post("/api/template")
def api_template_upload():
    with _pipeline_lock:
        if _pipeline_state["running"]:
            return jsonify({"error": "流水线运行中，请稍后再上传模板"}), 409

    upload = request.files.get("template")
    if upload is None or not upload.filename:
        return jsonify({"error": "请选择要上传的模板文件"}), 400

    original_name = Path(upload.filename).name
    if not original_name.lower().endswith(".xlsx"):
        return jsonify({"error": "仅支持 .xlsx 模板"}), 400

    content = upload.read()
    if not content:
        return jsonify({"error": "上传文件为空"}), 400
    try:
        _validate_xlsx(content)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    DEFAULT_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    backup = _backup_current_template()
    DEFAULT_TEMPLATE_PATH.write_bytes(content)
    meta = {
        "uploaded_as": original_name,
        "uploaded_at": beijing_strftime("%Y-%m-%d %H:%M:%S"),
        "size": len(content),
        "backup": backup,
        "safe_name": secure_filename(original_name) or "report_template.xlsx",
    }
    _template_meta_path().write_text(
        json.dumps(meta, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return jsonify({
        "message": f"已更新默认模板（来自 {upload.filename}）",
        "template": _template_info(),
        "backup": backup,
    })


@app.get("/api/report/latest")
def api_report_latest():
    path = _latest_report()
    if not path:
        return jsonify({"error": "暂无生成的报表"}), 404
    report = Path(path)
    return send_xlsx_file(report, display_name=report.name)


@app.post("/api/pipeline/run")
def api_pipeline_run():
    data = request.get_json(force=True, silent=True) or {}
    with _pipeline_lock:
        if _pipeline_state["running"]:
            return jsonify({"error": "任务正在运行中"}), 409

    settings = load_settings()
    year = int(data.get("year") or settings["year"])
    quarter = int(data.get("quarter") or settings["quarter"])
    notify = _parse_notify_flag(data, default=bool(settings.get("notify_feishu", True)))
    if notify is None:
        notify = bool(settings.get("notify_feishu", True))

    if data.get("save_settings", True):
        save_settings(year, quarter, notify_feishu=notify)

    skip_crawl = bool(data.get("skip_crawl"))
    threading.Thread(
        target=_run_pipeline_job,
        args=(year, quarter, skip_crawl),
        kwargs={"notify_feishu": bool(notify)},
        daemon=True,
    ).start()
    return jsonify({
        "started": True,
        "year": year,
        "quarter": quarter,
        "skip_crawl": skip_crawl,
        "notify_feishu": bool(notify),
    })


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default=os.getenv("CONSOLE_HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.getenv("CONSOLE_PORT", "9003")))
    args = parser.parse_args()
    print(f"区域报表自动化控制台 http://{args.host}:{args.port}")
    print(f"Session home: {get_session_home(PLUGIN_ROOT)}")
    app.run(host=args.host, port=args.port, debug=False, use_reloader=False)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
