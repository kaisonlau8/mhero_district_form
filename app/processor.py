from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ReportBuildError(ValueError):
    """Raised when uploaded files cannot be converted into a report."""


@dataclass(frozen=True)
class SourceSpec:
    key: str
    prefix: str
    target_sheet: str


SOURCE_SPECS: tuple[SourceSpec, ...] = (
    SourceSpec("stock", "门店备件库存导出", "备件库存明细"),
    SourceSpec("task", "保养提醒任务列表", "招揽实施率"),
    SourceSpec("first_service", "首保实施率", "首保"),
    SourceSpec("second_service", "二保实施率", "二保"),
    SourceSpec("new_policy", "新保投保率", "新保"),
    SourceSpec("renewal", "续保投保率", "续保"),
    SourceSpec("renewal_backlog", "去年同期交付未新保车辆", "续保"),
)

TEMPLATE_PREFIX = "区域各指标情况一览"
NUMERIC_PATTERN = re.compile(r"^-?\d+(?:\.\d+)?$")


def normalize_filename(filename: str) -> str:
    stem = Path(filename or "").stem
    return re.sub(r"\s+", "", stem)


def detect_file_role(filename: str) -> str | None:
    normalized = normalize_filename(filename)

    if TEMPLATE_PREFIX in normalized:
        return "template"

    for spec in sorted(SOURCE_SPECS, key=lambda item: len(item.prefix), reverse=True):
        if normalized.startswith(spec.prefix) or spec.prefix in normalized:
            return spec.key

    return None


def load_rows_from_workbook(content: bytes, filename: str) -> list[tuple]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:  # pragma: no cover - depends on upload damage
        raise ReportBuildError(f"无法读取文件：{filename}") from exc

    try:
        worksheet = workbook.active
        try:
            worksheet.reset_dimensions()
        except Exception:
            pass

        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        if not rows:
            raise ReportBuildError(f"文件没有可读取的数据：{filename}")
        return rows
    finally:
        workbook.close()


def parse_numeric_text(value: object) -> object:
    if value is None:
        return None

    if isinstance(value, (int, float, Decimal)):
        return value

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    if not NUMERIC_PATTERN.fullmatch(text):
        return value

    try:
        number = Decimal(text)
    except InvalidOperation:
        return value

    if number == number.to_integral():
        return int(number)

    return float(number)


def has_meaningful_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def normalize_stock_rows(rows: Iterable[tuple]) -> list[list[object]]:
    normalized_rows: list[list[object]] = []
    for row in rows:
        current = list(row)
        if len(current) < 21:
            current.extend([None] * (21 - len(current)))

        # 门店备件库存导出的 M 列需要写成数字，相关公式才会参与计算。
        current[12] = parse_numeric_text(current[12])
        normalized_rows.append(current)
    return normalized_rows


def normalize_first_service_rows(rows: Iterable[tuple]) -> list[list[object]]:
    normalized_rows: list[list[object]] = []
    for row in rows:
        current = list(row)
        if len(current) < 12:
            current.extend([None] * (12 - len(current)))

        # DMS 在超期首保和二次交付场景下，H 列可能不可靠。
        # 这里统一以 I 列“实际首保日期”是否有值作为最终判断依据。
        current[7] = "是" if has_meaningful_value(current[8]) else "否"
        normalized_rows.append(current)
    return normalized_rows


def normalize_backlog_rows(rows: Iterable[tuple]) -> list[list[object]]:
    normalized_rows: list[list[object]] = []
    for row in rows:
        current = list(row)
        if len(current) < 12:
            current.extend([None] * (12 - len(current)))

        # 去年同期交付未新保车辆缺少“保险到期时间”列，需要在 E/F 之间补空列。
        aligned = current[:5] + [None] + current[5:]
        normalized_rows.append(aligned)
    return normalized_rows


def clear_sheet_data(worksheet: Worksheet) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def write_rows(worksheet: Worksheet, rows: Iterable[Iterable[object]]) -> None:
    for row in rows:
        worksheet.append(list(row))


def prepare_template_workbook(template_content: bytes, template_name: str | None = None) -> Workbook:
    try:
        workbook = load_workbook(BytesIO(template_content), data_only=False)
    except Exception as exc:  # pragma: no cover - depends on upload damage
        raise ReportBuildError(f"模板无法打开：{template_name or '内置模板'}") from exc

    calc = workbook.calculation
    if calc is not None:
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True

    return workbook


def choose_output_name() -> str:
    date_suffix = datetime.now().strftime("%m%d")
    return f"{TEMPLATE_PREFIX}{date_suffix}.xlsx"


def build_report(
    uploaded_files: list[tuple[str, bytes]],
    template_content: bytes,
    template_name: str | None = None,
) -> tuple[bytes, str]:
    categorized: dict[str, tuple[str, bytes]] = {}
    embedded_template: tuple[str, bytes] | None = None

    for filename, content in uploaded_files:
        role = detect_file_role(filename)
        if role == "template":
            embedded_template = (filename, content)
            continue
        if role is None:
            continue
        categorized[role] = (filename, content)

    if embedded_template is not None and template_name is None:
        template_name, template_content = embedded_template

    missing = [spec.prefix for spec in SOURCE_SPECS if spec.key not in categorized]
    if missing:
        raise ReportBuildError("缺少以下文件：" + "、".join(missing))

    workbook = prepare_template_workbook(template_content, template_name)

    try:
        direct_sheet_sources = {
            "备件库存明细": "stock",
            "招揽实施率": "task",
            "首保": "first_service",
            "二保": "second_service",
            "新保": "new_policy",
        }

        for sheet_name, source_key in direct_sheet_sources.items():
            if sheet_name not in workbook.sheetnames:
                raise ReportBuildError(f"模板缺少工作表：{sheet_name}")

            _, source_content = categorized[source_key]
            source_rows = load_rows_from_workbook(source_content, categorized[source_key][0])
            data_rows = source_rows[1:]

            if source_key == "stock":
                data_rows = normalize_stock_rows(data_rows)
            elif source_key == "first_service":
                data_rows = normalize_first_service_rows(data_rows)

            sheet = workbook[sheet_name]
            clear_sheet_data(sheet)
            write_rows(sheet, data_rows)

        renewal_sheet_name = "续保"
        if renewal_sheet_name not in workbook.sheetnames:
            raise ReportBuildError(f"模板缺少工作表：{renewal_sheet_name}")

        renewal_rows = load_rows_from_workbook(categorized["renewal"][1], categorized["renewal"][0])[1:]
        backlog_rows = load_rows_from_workbook(
            categorized["renewal_backlog"][1],
            categorized["renewal_backlog"][0],
        )[1:]
        merged_rows = [list(row) for row in renewal_rows] + normalize_backlog_rows(backlog_rows)

        renewal_sheet = workbook[renewal_sheet_name]
        clear_sheet_data(renewal_sheet)
        write_rows(renewal_sheet, merged_rows)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue(), choose_output_name()
    finally:
        workbook.close()
