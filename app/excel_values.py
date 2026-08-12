#!/usr/bin/env python3
"""Materialize Excel formulas to plain values in pure Python (no Excel app)."""

from __future__ import annotations

import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[^'!]+)!)?(?P<col>\$?[A-Z]{1,3})(?P<row>\$?\d+)$",
    re.IGNORECASE,
)
COL_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[^'!]+)!)?(?P<c1>\$?[A-Z]{1,3}):(?P<c2>\$?[A-Z]{1,3})$",
    re.IGNORECASE,
)
RANGE_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[^'!]+)!)?(?P<c1>\$?[A-Z]{1,3})(?P<r1>\$?\d+):"
    r"(?P<c2>\$?[A-Z]{1,3})(?P<r2>\$?\d+)$",
    re.IGNORECASE,
)
STOCK_SUMIFS_RE = re.compile(
    r"^=SUMIFS\(备件库存明细!\$M:\$M,\s*备件库存明细!\$B:\$B,\s*\$C(\d+),"
    r"\s*备件库存明细!\$C:\$C,\s*([A-Z]{1,3})\$1\)$",
    re.IGNORECASE,
)


def _norm_sheet(name: str | None, default: str) -> str:
    if not name:
        return default
    name = name.strip()
    if name.startswith("'") and name.endswith("'"):
        name = name[1:-1]
    return name


def _col_idx(token: str) -> int:
    return column_index_from_string(token.replace("$", "").upper())


def _row_idx(token: str) -> int:
    return int(token.replace("$", ""))


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _excel_equal(a: Any, b: Any) -> bool:
    if a is None and (b is None or b == ""):
        return True
    if b is None and (a is None or a == ""):
        return True
    na, nb = _to_number(a), _to_number(b)
    if na is not None and nb is not None:
        return abs(na - nb) < 1e-9
    return str(a).strip() == str(b).strip()


def _match_criteria(value: Any, criteria: Any) -> bool:
    if isinstance(criteria, str):
        c = criteria.strip()
        if c == ">0":
            number = _to_number(value)
            return number is not None and number > 0
        m = re.match(r"^(>=|<=|<>|>|<)(.+)$", c)
        if m:
            op, raw = m.group(1), m.group(2)
            threshold = _to_number(raw)
            number = _to_number(value)
            if threshold is not None and number is not None:
                if op == ">=":
                    return number >= threshold - 1e-12
                if op == "<=":
                    return number <= threshold + 1e-12
                if op == ">":
                    return number > threshold
                if op == "<":
                    return number < threshold
                if op == "<>":
                    return abs(number - threshold) > 1e-12
            text = "" if value is None else str(value)
            if op == "<>":
                return text != raw
            return False
    return _excel_equal(value, criteria)


class WorkbookCalculator:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.wb = load_workbook(self.path, data_only=False)
        self.grids: dict[str, dict[int, dict[int, Any]]] = {}
        self.formulas: dict[tuple[str, int, int], str] = {}
        self.max_row: dict[str, int] = {}
        self.max_col: dict[str, int] = {}
        self.cache: dict[tuple[str, int, int], Any] = {}
        self.col_values: dict[tuple[str, int], list[Any]] = {}
        self._evaluating: set[tuple[str, int, int]] = set()
        self._load()

    def _load(self) -> None:
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            grid: dict[int, dict[int, Any]] = {}
            mr = ws.max_row or 1
            mc = ws.max_column or 1
            for r in range(1, mr + 1):
                row_map: dict[int, Any] = {}
                for c in range(1, mc + 1):
                    val = ws.cell(r, c).value
                    if isinstance(val, str) and val.startswith("="):
                        self.formulas[(name, r, c)] = val
                        row_map[c] = None
                    else:
                        row_map[c] = val
                grid[r] = row_map
            self.grids[name] = grid
            self.max_row[name] = mr
            self.max_col[name] = mc

    def _raw_cell(self, sheet: str, row: int, col: int) -> Any:
        return self.grids.get(sheet, {}).get(row, {}).get(col)

    def _get_col_values(self, sheet: str, col: int) -> list[Any]:
        key = (sheet, col)
        if key in self.col_values:
            return self.col_values[key]
        values = [self.get_value(sheet, r, col) for r in range(1, self.max_row.get(sheet, 1) + 1)]
        self.col_values[key] = values
        return values

    def get_value(self, sheet: str, row: int, col: int) -> Any:
        key = (sheet, row, col)
        if key in self.cache:
            return self.cache[key]
        if key in self.formulas:
            if key in self._evaluating:
                return None
            self._evaluating.add(key)
            try:
                result = self._eval_formula(sheet, self.formulas[key])
            finally:
                self._evaluating.discard(key)
            self.cache[key] = result
            col_key = (sheet, col)
            if col_key in self.col_values and 1 <= row <= len(self.col_values[col_key]):
                self.col_values[col_key][row - 1] = result
            return result
        return self._raw_cell(sheet, row, col)

    def _split_args(self, args_blob: str) -> list[str]:
        args: list[str] = []
        buf: list[str] = []
        depth = 0
        in_str = False
        for ch in args_blob:
            if ch == '"':
                in_str = not in_str
                buf.append(ch)
                continue
            if not in_str:
                if ch == "(":
                    depth += 1
                elif ch == ")":
                    depth -= 1
                elif ch == "," and depth == 0:
                    args.append("".join(buf).strip())
                    buf = []
                    continue
            buf.append(ch)
        if buf:
            args.append("".join(buf).strip())
        return args

    def _split_top_level(self, expr: str, sep: str) -> list[str]:
        parts: list[str] = []
        buf: list[str] = []
        depth = 0
        in_str = False
        i = 0
        while i < len(expr):
            ch = expr[i]
            if ch == '"':
                in_str = not in_str
                buf.append(ch)
                i += 1
                continue
            if not in_str:
                if ch == "(":
                    depth += 1
                elif ch == ")":
                    depth -= 1
                elif depth == 0 and expr.startswith(sep, i):
                    parts.append("".join(buf).strip())
                    buf = []
                    i += len(sep)
                    continue
            buf.append(ch)
            i += 1
        parts.append("".join(buf).strip())
        return parts

    def _eval_formula(self, current_sheet: str, formula: str) -> Any:
        return self._eval_expr(current_sheet, formula[1:] if formula.startswith("=") else formula)

    def _eval_expr(self, current_sheet: str, expr: str) -> Any:
        expr = expr.strip()
        if not expr:
            return None

        if len(expr) >= 2 and expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        if re.fullmatch(r"-?\d+(\.\d+)?%?", expr):
            return _to_number(expr)

        # function call first (before operator splits)
        upper = expr.upper()
        for fname in (
            "IFERROR",
            "COUNTIFS",
            "COUNTIF",
            "SUMIFS",
            "AVERAGE",
            "INDEX",
            "MATCH",
            "RANK",
            "SUM",
            "IF",
        ):
            prefix = fname + "("
            if upper.startswith(prefix) and expr.endswith(")"):
                # ensure matching paren for whole expr
                depth = 0
                ok = True
                in_str = False
                for i, ch in enumerate(expr):
                    if ch == '"':
                        in_str = not in_str
                    elif not in_str:
                        if ch == "(":
                            depth += 1
                        elif ch == ")":
                            depth -= 1
                            if depth == 0 and i != len(expr) - 1:
                                ok = False
                                break
                if ok and depth == 0:
                    args = self._split_args(expr[len(fname) + 1 : -1])
                    return self._call_func(current_sheet, fname, args)

        if "&" in expr:
            parts = self._split_top_level(expr, "&")
            if len(parts) > 1:
                out = []
                for part in parts:
                    val = self._eval_expr(current_sheet, part)
                    out.append("" if val is None else str(val))
                return "".join(out)

        for op in (">=", "<=", "<>", ">", "<"):
            parts = self._split_top_level(expr, op)
            if len(parts) == 2:
                left = self._eval_expr(current_sheet, parts[0])
                right = self._eval_expr(current_sheet, parts[1])
                if op == "<>":
                    return not _excel_equal(left, right)
                ln, rn = _to_number(left), _to_number(right)
                if ln is None or rn is None:
                    return False
                if op == ">=":
                    return ln >= rn
                if op == "<=":
                    return ln <= rn
                if op == ">":
                    return ln > rn
                if op == "<":
                    return ln < rn

        # Addition / subtraction (skip leading sign-only)
        parts = self._split_top_level(expr, "+")
        if len(parts) > 1:
            total = 0.0
            for part in parts:
                n = _to_number(self._eval_expr(current_sheet, part))
                total += 0.0 if n is None else n
            return total
        parts = self._split_top_level(expr, "-")
        if len(parts) > 1 and not expr.startswith("-"):
            first = _to_number(self._eval_expr(current_sheet, parts[0]))
            acc = 0.0 if first is None else first
            for part in parts[1:]:
                n = _to_number(self._eval_expr(current_sheet, part))
                acc -= 0.0 if n is None else n
            return acc

        parts = self._split_top_level(expr, "*")
        if len(parts) > 1:
            total = 1.0
            for part in parts:
                n = _to_number(self._eval_expr(current_sheet, part))
                total *= 0.0 if n is None else n
            return total
        parts = self._split_top_level(expr, "/")
        if len(parts) > 1:
            first = _to_number(self._eval_expr(current_sheet, parts[0]))
            if first is None:
                return None
            for part in parts[1:]:
                n = _to_number(self._eval_expr(current_sheet, part))
                if n is None or n == 0:
                    return None
                first /= n
            return first

        if expr.startswith("(") and expr.endswith(")"):
            return self._eval_expr(current_sheet, expr[1:-1])

        m = CELL_REF_RE.fullmatch(expr.replace(" ", ""))
        if m:
            sheet = _norm_sheet(m.group("sheet"), current_sheet)
            return self.get_value(sheet, _row_idx(m.group("row")), _col_idx(m.group("col")))

        raise ValueError(f"Unsupported expression: {expr}")

    def _resolve_range(self, current_sheet: str, token: str) -> list[Any]:
        token = token.strip()
        compact = token.replace(" ", "")
        m = RANGE_REF_RE.fullmatch(compact)
        if m:
            sheet = _norm_sheet(m.group("sheet"), current_sheet)
            c1, c2 = _col_idx(m.group("c1")), _col_idx(m.group("c2"))
            r1, r2 = _row_idx(m.group("r1")), _row_idx(m.group("r2"))
            vals: list[Any] = []
            for r in range(min(r1, r2), max(r1, r2) + 1):
                for c in range(min(c1, c2), max(c1, c2) + 1):
                    vals.append(self.get_value(sheet, r, c))
            return vals
        m = COL_REF_RE.fullmatch(compact)
        if m:
            sheet = _norm_sheet(m.group("sheet"), current_sheet)
            c1, c2 = _col_idx(m.group("c1")), _col_idx(m.group("c2"))
            if c1 != c2:
                raise ValueError(f"Only single-column refs supported: {token}")
            return self._get_col_values(sheet, c1)
        raise ValueError(f"Not a range: {token}")

    def _eval_criteria_arg(self, current_sheet: str, raw: str) -> Any:
        raw = raw.strip()
        if raw.startswith('"') and raw.endswith('"'):
            return raw[1:-1]
        return self._eval_expr(current_sheet, raw)

    def _call_func(self, current_sheet: str, fname: str, args: list[str]) -> Any:
        if fname == "IFERROR":
            try:
                val = self._eval_expr(current_sheet, args[0])
                if val is None:
                    return self._eval_expr(current_sheet, args[1])
                return val
            except Exception:
                return self._eval_expr(current_sheet, args[1])

        if fname == "IF":
            cond = self._eval_expr(current_sheet, args[0])
            return self._eval_expr(current_sheet, args[1] if cond else args[2])

        if fname == "SUM":
            total = 0.0
            for arg in args:
                seq = self._resolve_range(current_sheet, arg) if ":" in arg else [self._eval_expr(current_sheet, arg)]
                for v in seq:
                    n = _to_number(v)
                    if n is not None:
                        total += n
            return total

        if fname == "AVERAGE":
            nums = []
            for arg in args:
                seq = self._resolve_range(current_sheet, arg) if ":" in arg else [self._eval_expr(current_sheet, arg)]
                for v in seq:
                    n = _to_number(v)
                    if n is not None:
                        nums.append(n)
            return sum(nums) / len(nums) if nums else None

        if fname == "COUNTIF":
            values = self._resolve_range(current_sheet, args[0])
            criteria = self._eval_criteria_arg(current_sheet, args[1])
            return sum(1 for v in values if _match_criteria(v, criteria))

        if fname == "COUNTIFS":
            ranges = []
            criteria_list = []
            for i in range(0, len(args), 2):
                ranges.append(self._resolve_range(current_sheet, args[i]))
                criteria_list.append(self._eval_criteria_arg(current_sheet, args[i + 1]))
            n = min(len(r) for r in ranges)
            count = 0
            for idx in range(n):
                if all(_match_criteria(ranges[j][idx], criteria_list[j]) for j in range(len(ranges))):
                    count += 1
            return count

        if fname == "SUMIFS":
            sum_range = self._resolve_range(current_sheet, args[0])
            ranges = []
            criteria_list = []
            for i in range(1, len(args), 2):
                ranges.append(self._resolve_range(current_sheet, args[i]))
                criteria_list.append(self._eval_criteria_arg(current_sheet, args[i + 1]))
            n = min([len(sum_range)] + [len(r) for r in ranges])
            total = 0.0
            for idx in range(n):
                if all(_match_criteria(ranges[j][idx], criteria_list[j]) for j in range(len(ranges))):
                    number = _to_number(sum_range[idx])
                    if number is not None:
                        total += number
            return total

        if fname == "RANK":
            number = _to_number(self._eval_expr(current_sheet, args[0]))
            values = self._resolve_range(current_sheet, args[1])
            nums = [n for n in (_to_number(v) for v in values) if n is not None]
            if number is None:
                return None
            return 1 + sum(1 for n in nums if n > number)

        if fname == "MATCH":
            lookup = self._eval_expr(current_sheet, args[0])
            values = self._resolve_range(current_sheet, args[1])
            for i, v in enumerate(values, start=1):
                if _excel_equal(v, lookup):
                    return i
            return None

        if fname == "INDEX":
            values = self._resolve_range(current_sheet, args[0])
            row_num = _to_number(self._eval_expr(current_sheet, args[1]))
            if row_num is None:
                return None
            idx = int(row_num)
            if 1 <= idx <= len(values):
                return values[idx - 1]
            return None

        raise ValueError(f"Unsupported function: {fname}")

    def _precompute_stock_sumifs(self) -> None:
        stock = "备件库存明细"
        if stock not in self.grids:
            return
        idx: dict[tuple[str, str], float] = defaultdict(float)
        for r in range(2, self.max_row[stock] + 1):
            b = self._raw_cell(stock, r, 2)
            c = self._raw_cell(stock, r, 3)
            m = _to_number(self._raw_cell(stock, r, 13))
            if b is None or c is None or m is None:
                continue
            idx[(str(b).strip(), str(c).strip())] += m

        sheet = "常备件备库率"
        for (sh, row, col), formula in self.formulas.items():
            if sh != sheet:
                continue
            m = STOCK_SUMIFS_RE.match(formula.replace(" ", ""))
            if not m:
                continue
            store_row = int(m.group(1))
            part_col = _col_idx(m.group(2))
            store = self._raw_cell(sheet, store_row, 3)
            part = self._raw_cell(sheet, 1, part_col)
            if store is None or part is None:
                self.cache[(sh, row, col)] = 0
            else:
                self.cache[(sh, row, col)] = float(
                    idx.get((str(store).strip(), str(part).strip()), 0.0)
                )

    def materialize(self) -> None:
        self._precompute_stock_sumifs()
        for sheet, row, col in list(self.formulas):
            self.get_value(sheet, row, col)
        for (sheet, row, col), _formula in self.formulas.items():
            self.wb[sheet].cell(row, col).value = self.cache.get((sheet, row, col))

    def save(self, destination: Path) -> None:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(destination)
        self.wb.close()


def flatten_formulas_to_values(
    source: Path,
    destination: Path | None = None,
    *,
    visible: bool = False,
) -> Path:
    """Calculate formulas in pure Python and save a values-only workbook."""
    del visible
    source = Path(source).resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    if destination is None:
        destination = source
    else:
        destination = Path(destination).resolve()
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination != source:
            shutil.copy2(source, destination)

    calc = WorkbookCalculator(destination)
    calc.materialize()
    calc.save(destination)
    return destination


def flatten_bytes(content: bytes, *, suffix: str = ".xlsx") -> bytes:
    with tempfile.TemporaryDirectory(prefix="district-form-flat-") as tmp:
        dst = Path(tmp) / f"values{suffix}"
        dst.write_bytes(content)
        flatten_formulas_to_values(dst)
        return dst.read_bytes()
